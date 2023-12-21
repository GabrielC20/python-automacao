from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
# 1 lê pasta de trabalho e planilha

wb = load_workbook("data/pivot_table.xlsx")
sheet = wb ["Relatorio"]

# 2 referências das linhas e colunas 

min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# 3-adicionando dados e categorias no gráfico

barchart = BarChart()

data = Reference(
    sheet,
    min_col=min_column + 1,
    max_col=max_column,
    min_row=min_row,
    max_row=max_row
)

categores = Reference(
    sheet, 
    min_col=min_column,
    max_col=min_column,
    min_row=min_row + 1,
    max_row=max_row
)

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categores)


#4-Criando o Gráfico

sheet.add_chart(barchart, "B10")
barchart.title = "Vendas por Fabricante"
barchart.style = 2

# 5- Salvando o workbook

wb.save("data/barchat.xlsx")
