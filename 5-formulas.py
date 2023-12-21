from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
# 1 lê pasta de trabalho e planilha

wb = load_workbook("data/barchat.xlsx")
sheet = wb ["Relatorio"]

# 2 referências das linhas e colunas 

min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# 3-inclúindo fórmulas

#sheet["B6"] = "=SUM(B2:B5)"
#sheet["B6"].style = "Currency"

for i in range(min_column+1, max_column+1):
    letter = get_column_letter(i)
    
    # Correção na fórmula
    sheet[f"{letter}{max_row+1}"] = f"=SUM({letter}{min_row+1}:{letter}{max_row})"
    sheet[f"{letter}{max_row+1}"].style = "Currency"

wb.save("test.xlsx")
